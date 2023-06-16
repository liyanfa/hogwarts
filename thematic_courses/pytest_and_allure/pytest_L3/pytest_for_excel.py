# -*- coding: utf-8 -*-
# @Time    : 2023/6/13 20:47
# @Author  : yanfa
# @user   : yanfa 
# @File    : pytest_for_excel.py
# @remark:pytest结合数据驱动excel

""""""
"""一、读取excel文件
第三方库：xlrd xlwings pandas
openpyxl：官方文档https://openpyxl.readthedocs.io/en/stable/"""

"""二、openpyxl库的安装
安装：pip install openpyxl
导入：import openpyxl"""

"""三、openpyxl库的操作
读取工作簿：load_workbook()
读取工作表：
读取单元格：
"""
import openpyxl

# 读取工作簿
file_path = "pytest_for_excel/data/data.xlsx"
book = openpyxl.load_workbook(file_path)

# 读取工作表
sheet = book.active

# 读取单个单元格，获取单元格的值用value方法
cell_a1 = sheet['A1'].value
cell_a3 = sheet.cell(column=1, row=3).value  # A3
print(cell_a1)
print(cell_a3)

# 读取多个连续单元格
cells = sheet["A1":"C3"]
print(cells)
